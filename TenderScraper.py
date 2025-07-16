import pandas as pd
import logging
import time
import os
import json
from datetime import datetime
from typing import List, Dict, Any, Set
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    StaleElementReferenceException, 
    ElementClickInterceptedException,
    TimeoutException,
    NoSuchElementException
)

from ConfigManager import ConfigManager
from Utils import Utils

class TenderScraper:
    """
    Main scraper class for extracting tender data from eTenders website.
    
    This class handles the complete scraping workflow including browser management,
    data extraction, error handling, duplicate prevention, and result export.
    """
    
    def __init__(self, configPath: str = "config.json"):
        """
        Initialize the TenderScraper with configuration.
        
        Args:
            configPath (str): Path to the configuration file
        """
        # Setup logging first (before ConfigManager which tries to log)
        self.setupLogging(configPath)
        
        self.configManager = ConfigManager(configPath)
        self.utils = Utils()
        self.driver = None
        self.tenderData = []
        self.processedTenders = set()  # Track processed tenders to prevent duplicates
        self.reportDate = datetime.now().strftime("%Y/%m/%d")
        
        # Get configuration
        # These could be simplified to direct access: self.configManager.config.get('scraping', {})
        # But using getter methods for possible future improvements.
        self.scrapingConfig = self.configManager.getScrapingConfig()
        self.browserConfig = self.configManager.getBrowserConfig()
        self.timingConfig = self.configManager.getTimingConfig()
        self.retryConfig = self.configManager.getRetryConfig()
        self.outputConfig = self.configManager.getOutputConfig()
        
        logging.info("TenderScraper initialized successfully")
    
    def setupLogging(self, configPath: str) -> None:
        """
        Setup logging configuration for the scraper.
        
        Args:
            configPath (str): Path to the configuration file
        """
        # Load config temporarily to get logging settings
        try:
            with open(configPath, 'r') as file:
                config = json.load(file)
            loggingConfig = config.get('logging', {})
        except Exception:
            # Fallback to default logging config
            loggingConfig = {'level': 'INFO', 'file': 'logs/scraper.log'}
        
        logging.basicConfig(
            level=getattr(logging, loggingConfig.get('level', 'INFO')),
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.FileHandler(loggingConfig.get('file', 'logs/scraper.log')),
                logging.StreamHandler()
            ]
        )
    
    def setupBrowser(self) -> None:
        """
        Setup and configure the Chrome browser for scraping.
        """
        chromeOptions = Options()
        
        if self.browserConfig.get('headless', False):
            chromeOptions.add_argument("--headless=new")
        
        if self.browserConfig.get('maximized', True):
            chromeOptions.add_argument("--start-maximized")
        
        if self.browserConfig.get('disableExtensions', True):
            chromeOptions.add_argument("--disable-extensions")
        
        if self.browserConfig.get('disableInfobars', True):
            chromeOptions.add_argument("--disable-infobars")
        
        # Additional stability options
        chromeOptions.add_argument("--no-sandbox")
        chromeOptions.add_argument("--disable-dev-shm-usage")
        chromeOptions.add_argument("--disable-gpu")
        
        self.driver = webdriver.Chrome(options=chromeOptions)
        logging.info("Browser setup completed")
    
    def navigateToPage(self) -> None:
        """
        Navigate to the eTenders opportunities page and wait for it to load.
        """
        url = self.scrapingConfig.get('url')
        logging.info(f"Navigating to: {url}")
        
        self.driver.get(url)
        time.sleep(self.timingConfig.get('pageLoadWait', 7))
        
        # Remove any modal popups
        self.removeModalPopups()
        
        logging.info("Page loaded successfully")
    
    def removeModalPopups(self) -> None:
        """
        Remove any modal popups that might interfere with scraping.
        """
        try:
            self.driver.execute_script("""
                let modals = document.querySelectorAll('.modal.show');
                modals.forEach(m => m.parentNode.removeChild(m));
                let backdrop = document.querySelector('.modal-backdrop');
                if (backdrop) backdrop.parentNode.removeChild(backdrop);
                document.body.classList.remove('modal-open');
            """)
            time.sleep(self.timingConfig.get('modalRemovalWait', 1))
            logging.debug("Modal popups removed")
        except Exception as e:
            logging.warning(f"Failed to remove modal popups: {e}")
    
    def getValueByLabel(self, label: str) -> str:
        """
        Extract value from expanded tender details by label.
        
        Args:
            label (str): Label text to search for
            
        Returns:
            str: Extracted value or empty string if not found
        """
        try:
            element = self.driver.find_element(By.XPATH, f"//b[contains(text(), '{label}')]")
            value = element.find_element(By.XPATH, "./ancestor::td[1]/following-sibling::td[1]").text
            return self.utils.cleanText(value)
        except NoSuchElementException:
            return ""
        except Exception as e:
            logging.warning(f"Error extracting value for label '{label}': {e}")
            return ""
    
    def extractDocumentLinks(self) -> List[str]:
        """
        Extract document download links from the expanded tender details.
        
        Returns:
            List[str]: List of document download URLs
        """
        docLinks = []
        try:
            docSection = self.driver.find_element(By.XPATH, "//b[contains(text(),'TENDER DOCUMENTS')]/ancestor::table[1]")
            aTags = docSection.find_elements(By.TAG_NAME, "a")
            
            for a in aTags:
                href = a.get_attribute("href")
                if href and "Download" in href:
                    docLinks.append(href)
            
            logging.debug(f"Found {len(docLinks)} document links")
        except NoSuchElementException:
            logging.debug("No document section found")
        except Exception as e:
            logging.warning(f"Error extracting document links: {e}")
        
        return docLinks
    
    def processTenderRow(self, row, rowIndex: int, pageNum: int) -> bool:
        """
        Process a single tender row and extract all relevant data.
        
        Args:
            row: Selenium WebElement representing the tender row
            rowIndex (int): Index of the row on current page
            pageNum (int): Current page number
            
        Returns:
            bool: True if tender was successfully processed, False otherwise
        """
        try:
            # Get table cells
            tds = row.find_elements(By.TAG_NAME, "td")
            if len(tds) < 5:
                logging.debug(f"Row {rowIndex + 1} on page {pageNum} has insufficient columns")
                return False
            
            # Extract basic information
            category = self.utils.cleanText(tds[1].text)
            tenderDescription = self.utils.cleanText(tds[2].text)
            esubRaw = self.utils.cleanText(tds[3].text)
            advertisedStr = self.utils.cleanText(tds[4].text)
            
            # Parse publication date
            try:
                advertisedDt = datetime.strptime(advertisedStr, "%d/%m/%Y")
            except ValueError:
                logging.warning(f"Could not parse publication date: {advertisedStr}")
                return False
            
            # Check date range
            dateFrom = datetime.strptime(self.scrapingConfig['dateFrom'], "%Y-%m-%d")
            dateTo = datetime.strptime(self.scrapingConfig['dateTo'], "%Y-%m-%d")
            
            if not (dateFrom <= advertisedDt <= dateTo):
                if advertisedDt < dateFrom:
                    logging.info(f"Reached tenders older than date range, stopping")
                    return "STOP_SCRAPING"
                return False
            
            # Process e-submission status
            if "✔" in esubRaw or "tick" in esubRaw.lower():
                esubmission = "Yes"
            elif "x" in esubRaw.lower():
                esubmission = "No"
            elif esubRaw == "":
                esubmission = "No"
            else:
                esubmission = esubRaw
            
            # Expand row to get detailed information
            expandBtn = tds[0]
            try:
                expandBtn.click()
                time.sleep(self.timingConfig.get('expandRowWait', 3))
            except ElementClickInterceptedException:
                time.sleep(2)
                expandBtn.click()
                time.sleep(self.timingConfig.get('expandRowWait', 3))
            
            # Extract detailed information
            tenderInfo = {
                "REPORT_DATE": self.reportDate,
                "RECORD_ID": None,  # Will be filled later
                "TENDER_ID": self.getValueByLabel("Tender Number:"),
                "PUBLICATION_DATE": advertisedDt.strftime("%Y/%m/%d"),
                "CLOSING_DATE": "",
                "CLOSING_TIME": "",
                "TENDER_TYPE": self.getValueByLabel("Tender Type:"),
                "TENDER_DESCRIPTION": tenderDescription,
                "BRIEFING_SESSION_VENUE": self.getValueByLabel("Briefing Venue"),
                "TENDER_SOURCE": "ETENDERS.GOV.ZA",
                "DEPARTMENT": self.getValueByLabel("Organ Of State:"),
                "PROVINCE": self.getValueByLabel("Province:"),
                "ESUBMISSION": esubmission,
                "CATEGORY": category,
                "IS_THERE_A_BRIEFING_SESSION": self.getValueByLabel("Is there a briefing session?"),
                "BRIEFING_DATE": "",
                "COMPULSORY_BRIEFING": self.getValueByLabel("Is it compulsory?"),
                "LINK": "",
                "SOE": "",
                "COST_OF_SALES_ESTIMATE": "",
                "CAPABILITY_AVAILABLE": "",
                "CAPABILITY_GROUP": "",
                "REQUIREMENTS": ""
            }
            
            # Parse closing date and time
            closingDateRaw = self.getValueByLabel("Closing Date:")
            closingDate, closingTime = self.utils.parseClosingDateTime(closingDateRaw)
            tenderInfo["CLOSING_DATE"] = closingDate
            tenderInfo["CLOSING_TIME"] = closingTime
            
            # Parse briefing date
            briefingDateRaw = self.getValueByLabel("Briefing Date and Time")
            tenderInfo["BRIEFING_DATE"] = self.utils.parseDayMonthYear(briefingDateRaw)
            
            # Extract document links
            docLinks = self.extractDocumentLinks()
            tenderInfo["LINK"] = docLinks[0] if docLinks else ""
            
            # Check for duplicates
            if self.utils.isDuplicate(tenderInfo, self.processedTenders):
                logging.info(f"Duplicate tender found: {tenderInfo['TENDER_DESCRIPTION']}")
                # Collapse row and continue
                try:
                    expandBtn.click()
                    time.sleep(self.timingConfig.get('collapseRowWait', 2))
                except Exception:
                    pass
                return False
            
            # Clean and validate data
            tenderInfo = self.utils.cleanTenderData(tenderInfo)
            
            if not self.utils.validateTenderData(tenderInfo):
                logging.warning(f"Invalid tender data for: {tenderInfo['TENDER_DESCRIPTION']}")
                return False
            
            # Add to processed set and data list
            uniqueKey = self.utils.generateUniqueKey(tenderInfo)
            self.processedTenders.add(uniqueKey)
            self.tenderData.append(tenderInfo)
            
            logging.info(f"Tender found: {tenderInfo['TENDER_DESCRIPTION']}, Advertised: {tenderInfo['PUBLICATION_DATE']}")
            
            # Collapse row
            try:
                expandBtn.click()
                time.sleep(self.timingConfig.get('collapseRowWait', 2))
            except Exception:
                pass
            
            return True
            
        except StaleElementReferenceException as e:
            logging.warning(f"Stale element on row {rowIndex + 1}, page {pageNum}: {e}")
            return False
        except Exception as e:
            logging.error(f"Error processing row {rowIndex + 1} on page {pageNum}: {e}")
            return False
    
    def scrapePage(self, pageNum: int) -> bool:
        """
        Scrape all tender rows from the current page.
        
        Args:
            pageNum (int): Current page number
            
        Returns:
            bool: True if should continue to next page, False if should stop
        """
        logging.info(f"--- Scraping page {pageNum} ---")
        
        rowIndex = 0
        maxRetries = self.retryConfig.get('maxRetries', 3)
        
        while True:
            # Always refetch rows to avoid stale element issues
            try:
                rows = self.driver.find_elements(By.XPATH, "//table[contains(@class, 'dataTable')]/tbody/tr")
                mainRows = [row for row in rows if not row.get_attribute("innerHTML").startswith('<td colspan')]
                
                if rowIndex >= len(mainRows):
                    break
                
                # Process row with retry logic
                retryCount = 0
                while retryCount <= maxRetries:
                    try:
                        result = self.processTenderRow(mainRows[rowIndex], rowIndex, pageNum)
                        
                        if result == "STOP_SCRAPING":
                            return False
                        elif result:
                            break  # Success, move to next row
                        else:
                            break  # Failed but not retryable
                            
                    except StaleElementReferenceException:
                        retryCount += 1
                        if retryCount <= maxRetries:
                            logging.warning(f"Stale element on row {rowIndex + 1}, page {pageNum}. Retrying ({retryCount}/{maxRetries})...")
                            time.sleep(self.timingConfig.get('retryDelay', 2.5))
                            # Refetch rows after stale element
                            rows = self.driver.find_elements(By.XPATH, "//table[contains(@class, 'dataTable')]/tbody/tr")
                            mainRows = [row for row in rows if not row.get_attribute("innerHTML").startswith('<td colspan')]
                            continue
                        else:
                            logging.error(f"Failed row {rowIndex + 1} after {maxRetries} retries. Skipping.")
                            break
                
                rowIndex += 1
                
            except Exception as e:
                logging.error(f"Error processing page {pageNum}: {e}")
                break
        
        return True
    
    def goToNextPage(self) -> bool:
        """
        Navigate to the next page of results.
        
        Returns:
            bool: True if successfully moved to next page, False if no more pages
        """
        try:
            nextBtn = self.driver.find_element(By.ID, "tendeList_next")
            
            if "disabled" in nextBtn.get_attribute("class"):
                logging.info("Next button is disabled, no more pages")
                return False
            
            nextBtn.click()
            time.sleep(self.timingConfig.get('nextPageWait', 4))
            return True
            
        except NoSuchElementException:
            logging.info("Next button not found, no more pages")
            return False
        except Exception as e:
            logging.error(f"Error navigating to next page: {e}")
            return False
    
    def exportToExcel(self) -> None:
        """
        Export scraped data to Excel files (date-specific and cumulative) with proper formatting.
        """
        if not self.tenderData:
            logging.warning("No tender data to export")
            return
        
        # Define column order
        finalColumns = [
            "REPORT_DATE", "RECORD_ID", "TENDER_ID", "PUBLICATION_DATE",
            "CLOSING_DATE", "CLOSING_TIME", "TENDER_TYPE", "TENDER_DESCRIPTION",
            "TENDER_SOURCE", "DEPARTMENT", "PROVINCE",
            "ESUBMISSION", "CATEGORY", "IS_THERE_A_BRIEFING_SESSION",
            "BRIEFING_DATE", "COMPULSORY_BRIEFING", "BRIEFING_SESSION_VENUE", "LINK", "SOE",
            "COST_OF_SALES_ESTIMATE", "CAPABILITY_AVAILABLE", "CAPABILITY_GROUP", "REQUIREMENTS"
        ]
        
        # Generate filenames
        # Use dateTo for date-specific file to prevent overwriting
        dateToStr = self.scrapingConfig.get('dateTo', '').replace('-', '_')
        dateSpecificFile = self.outputConfig.get('dateSpecificFile', 'tenders_{date}.xlsx').format(date=dateToStr)
        cumulativeFile = self.outputConfig.get('cumulativeFile', 'master_tenders.xlsx')
        
        # Prepare date-specific data (preserve scraping order: first scraped = highest ID)
        dateSpecificData = []
        for idx, tender in enumerate(self.tenderData, 1):
            tenderCopy = tender.copy()
            tenderCopy["RECORD_ID"] = len(self.tenderData) - idx + 1  # Reverse order
            dateSpecificData.append(tenderCopy)
        
        dfDateSpecific = pd.DataFrame(dateSpecificData)[finalColumns]
        
        # Save date-specific file with proper formatting
        self._saveExcelWithFormatting(dfDateSpecific, dateSpecificFile)
        logging.info(f"Saved date-specific results to {dateSpecificFile}")
        
        # Handle cumulative file - load existing data and append new data
        cumulativeData = []
        
        # Try to load existing cumulative file
        try:
            if os.path.exists(cumulativeFile):
                existingDf = pd.read_excel(cumulativeFile)
                cumulativeData = existingDf.to_dict('records')
                logging.info(f"Loaded {len(cumulativeData)} existing records from {cumulativeFile}")
            else:
                logging.info(f"Creating new cumulative file: {cumulativeFile}")
        except Exception as e:
            logging.warning(f"Error loading existing cumulative file: {e}")
        
        # Add new tenders to cumulative data (at the top)
        existingTenderIds = {tender.get('TENDER_ID', '') for tender in cumulativeData}
        newTendersAdded = 0
        
        for tender in self.tenderData:
            tenderId = tender.get('TENDER_ID', '')
            if tenderId not in existingTenderIds:
                cumulativeData.insert(0, tender)  # Insert at the beginning (top)
                newTendersAdded += 1
            else:
                logging.debug(f"Skipping duplicate tender: {tenderId}")
        
        # Add record IDs to cumulative data (reverse order: newest = 1, oldest = highest number)
        for idx, tender in enumerate(cumulativeData, 1):
            tender["RECORD_ID"] = len(cumulativeData) - idx + 1
        
        # Save cumulative file with proper formatting
        if cumulativeData:
            dfCumulative = pd.DataFrame(cumulativeData)[finalColumns]
            self._saveExcelWithFormatting(dfCumulative, cumulativeFile)
            logging.info(f"Saved cumulative results to {cumulativeFile} ({newTendersAdded} new tenders added)")
        else:
            logging.warning("No data to save in cumulative file")
        
        logging.info(f"Total tenders exported: {len(self.tenderData)} (date-specific), {len(cumulativeData)} (cumulative)")
    
    def _saveExcelWithFormatting(self, df: pd.DataFrame, filename: str) -> None:
        """
        Save DataFrame to Excel with proper formatting for dates and links.
        
        Args:
            df (pd.DataFrame): DataFrame to save
            filename (str): Output filename
        """
        try:
            # First save with pandas to get the basic structure
            df.to_excel(filename, index=False, engine='openpyxl')
            
            # Now use openpyxl to apply proper formatting
            from openpyxl import load_workbook
            from openpyxl.styles import NamedStyle
            from openpyxl.utils.dataframe import dataframe_to_rows
            from datetime import datetime
            
            # Load the workbook
            wb = load_workbook(filename)
            ws = wb.active
            
            # Define date style
            date_style = NamedStyle(name="date_style")
            date_style.number_format = "YYYY/MM/DD"
            
            # Get column indices for date columns
            date_columns = ['REPORT_DATE', 'PUBLICATION_DATE', 'CLOSING_DATE', 'BRIEFING_DATE']
            date_col_indices = []
            
            for col_name in date_columns:
                if col_name in df.columns:
                    col_idx = df.columns.get_loc(col_name) + 1  # +1 because Excel columns are 1-indexed
                    date_col_indices.append(col_idx)
            
            # Apply date formatting to date columns
            for col_idx in date_col_indices:
                for row_idx in range(2, len(df) + 2):  # Start from row 2 (skip header)
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value:
                        try:
                            # Convert string date to datetime object
                            if isinstance(cell.value, str):
                                date_obj = datetime.strptime(cell.value, "%Y/%m/%d")
                                cell.value = date_obj
                            cell.style = date_style
                        except ValueError:
                            # If date parsing fails, keep as string
                            pass
            
            # Handle hyperlinks in LINK column
            link_col_idx = None
            if 'LINK' in df.columns:
                link_col_idx = df.columns.get_loc('LINK') + 1
                
                for row_idx in range(2, len(df) + 2):  # Start from row 2 (skip header)
                    cell = ws.cell(row=row_idx, column=link_col_idx)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('http'):
                        # Create hyperlink
                        cell.hyperlink = cell.value
                        cell.style = "Hyperlink"
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save the formatted workbook
            wb.save(filename)
            wb.close()
            
        except ImportError:
            # Fallback to basic pandas export if openpyxl is not available
            logging.warning("openpyxl not available, using basic Excel export")
            df.to_excel(filename, index=False)
        except Exception as e:
            logging.error(f"Error formatting Excel file {filename}: {e}")
            # Fallback to basic pandas export
            df.to_excel(filename, index=False)
    
    def run(self) -> None:
        """
        Run the complete scraping process.
        """
        try:
            logging.info("Starting tender scraping process")
            
            # Setup browser
            self.setupBrowser()
            
            # Navigate to page
            self.navigateToPage()
            
            # Scrape all pages
            pageNum = 1
            while True:
                shouldContinue = self.scrapePage(pageNum)
                if not shouldContinue:
                    break
                
                if not self.goToNextPage():
                    break
                
                pageNum += 1
            
            # Export results
            self.exportToExcel()
            
            logging.info("Scraping process completed successfully")
            
        except Exception as e:
            logging.error(f"Error during scraping process: {e}")
            raise
        finally:
            self.cleanup()
    
    def cleanup(self) -> None:
        """
        Clean up resources and close browser.
        """
        if self.driver:
            try:
                self.driver.quit()
                logging.info("Browser closed successfully")
            except Exception as e:
                logging.warning(f"Error closing browser: {e}") 